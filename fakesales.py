from faker import Faker
from openpyxl import Workbook
import random

# Initialize Faker
fake = Faker()

# Initialize Excel Workbook and Sheet
wb = Workbook()
ws = wb.active
ws.title = "Fake Sales Data"

# Add headers to the Excel sheet
headers = ["Customer Name", "Product", "Purchase Date", "Amount"]
for col_num, header in enumerate(headers, 1):
    col_letter = ws.cell(row=1, column=col_num)
    col_letter.value = header

# Sample product list
products = ["Laptop", "Smartphone", "TV", "Headphones", "Camera"]

# Generate fake sales data and populate the Excel sheet
for row_num in range(2, 102):  # 100 rows of data
    ws.cell(row=row_num, column=1).value = fake.name()
    ws.cell(row=row_num, column=2).value = random.choice(products)
    ws.cell(row=row_num, column=3).value = fake.date_between(start_date='-1y', end_date='today').strftime('%Y-%m-%d')
    ws.cell(row=row_num, column=4).value = round(random.uniform(50.5, 999.9), 2)

# Save the workbook
wb.save("fake_sales_data.xlsx")

print("Fake sales data generated and saved to 'fake_sales_data.xlsx'")


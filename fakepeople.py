from faker import Faker
from openpyxl import Workbook

# Initialize Faker
fake = Faker()

# Initialize Excel Workbook and Sheet
wb = Workbook()
ws = wb.active
ws.title = "Fake People Data"

# Add headers to the Excel sheet
headers = ["Name", "Email", "Address", "Phone Number", "Income"]
for col_num, header in enumerate(headers, 1):
    col_letter = ws.cell(row=1, column=col_num)
    col_letter.value = header

# Generate fake data and populate the Excel sheet
for row_num in range(2, 102):  # 100 rows of data
    ws.cell(row=row_num, column=1).value = fake.name()
    ws.cell(row=row_num, column=2).value = fake.email()
    ws.cell(row=row_num, column=3).value = fake.address()
    ws.cell(row=row_num, column=4).value = fake.phone_number()
    ws.cell(row=row_num, column=5).value = fake.random_int(min=30000, max=120000)

# Save the workbook
wb.save("fake_people_data.xlsx")

print("Fake data generated and saved to 'fake_people_data.xlsx'")

